from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
import re
import unicodedata
from typing import Callable

import openpyxl
import xlsxwriter
from openpyxl.utils import column_index_from_string, get_column_letter

from core.diff_engine import TextDiffer
from core.models import ChunkType, ParseError, ParsedDocument, Segment, UnsupportedFormatError
from core.registry import ParserRegistry


STATUS_APPLIED = "APPLIED"
STATUS_NOT_APPLIED = "NOT APPLIED"
STATUS_CANNOT_VERIFY = "CANNOT VERIFY"
STATUS_NOT_APPLICABLE = "NOT APPLICABLE"


REQUIRED_MAPPING_FIELDS = ("source_column", "original_column", "qa_mark_column")
SUPPORTED_FINAL_EXTENSIONS = {".xliff", ".xlf", ".sdlxliff", ".mqxliff"}

NBSP_CHARS = {
    "\u00A0",
    "\u202F",
    "\u2007",
}
APOSTROPHE_CHARS = {
    "\u2019",
    "\u2018",
    "\u02BC",
    "\u0060",
    "\u00B4",
    "\u2032",
}


@dataclass
class QAColumnInfo:
    column_index: int
    column_letter: str
    header: str

    def display_name(self) -> str:
        return f"{self.column_letter}: {self.header}"


@dataclass
class QAColumnMapping:
    source_column: str | None = None
    original_column: str | None = None
    revised_column: str | None = None
    qa_mark_column: str | None = None
    segment_id_column: str | None = None
    filename_column: str | None = None

    def is_complete(self) -> bool:
        return all(getattr(self, field_name) for field_name in REQUIRED_MAPPING_FIELDS)

    def to_dict(self) -> dict[str, str | None]:
        return {
            "source_column": self.source_column,
            "original_column": self.original_column,
            "revised_column": self.revised_column,
            "qa_mark_column": self.qa_mark_column,
            "segment_id_column": self.segment_id_column,
            "filename_column": self.filename_column,
        }

    @classmethod
    def from_dict(cls, payload: dict[str, str | None]) -> "QAColumnMapping":
        return cls(
            source_column=payload.get("source_column"),
            original_column=payload.get("original_column"),
            revised_column=payload.get("revised_column"),
            qa_mark_column=payload.get("qa_mark_column"),
            segment_id_column=payload.get("segment_id_column"),
            filename_column=payload.get("filename_column"),
        )


@dataclass
class QASheetConfig:
    report_path: str
    sheet_name: str
    header_row: int
    columns: list[QAColumnInfo]
    mapping: QAColumnMapping
    notes: list[str] = field(default_factory=list)

    def display_name(self) -> str:
        return f"{Path(self.report_path).name} :: {self.sheet_name}"

    def to_dict(self) -> dict:
        return {
            "report_path": self.report_path,
            "sheet_name": self.sheet_name,
            "header_row": self.header_row,
            "columns": [
                {
                    "column_index": item.column_index,
                    "column_letter": item.column_letter,
                    "header": item.header,
                }
                for item in self.columns
            ],
            "mapping": self.mapping.to_dict(),
            "notes": list(self.notes),
        }

    @classmethod
    def from_dict(cls, payload: dict) -> "QASheetConfig":
        return cls(
            report_path=str(payload.get("report_path", "")),
            sheet_name=str(payload.get("sheet_name", "")),
            header_row=int(payload.get("header_row", 1)),
            columns=[
                QAColumnInfo(
                    column_index=int(item["column_index"]),
                    column_letter=str(item["column_letter"]),
                    header=str(item["header"]),
                )
                for item in payload.get("columns", [])
            ],
            mapping=QAColumnMapping.from_dict(payload.get("mapping", {})),
            notes=[str(item) for item in payload.get("notes", [])],
        )


@dataclass
class QAScanResult:
    sheet_configs: list[QASheetConfig]
    warnings: list[str]


@dataclass
class QAVerificationRow:
    source: str
    original_translation: str
    revised_translation: str
    final_translation: str
    expected_file_name: str
    matched_file_name: str
    matched_file_path: str
    qa_mark: str
    verification_status: str
    matched_segment_id: str
    reason: str
    report_file: str
    sheet_name: str
    row_number: int


@dataclass
class QAFileSummary:
    file_name: str
    file_path: str
    qa_rows: int = 0
    applied: int = 0
    not_applied: int = 0
    cannot_verify: int = 0
    not_applicable: int = 0


@dataclass
class QAVerificationResult:
    rows: list[QAVerificationRow]
    status_counts: dict[str, int]
    warnings: list[str]
    total_rows: int
    timestamp: datetime
    file_summaries: list[QAFileSummary] = field(default_factory=list)


@dataclass
class _QARowInput:
    source: str
    original_translation: str
    revised_translation: str
    qa_mark_raw: str
    segment_id: str
    expected_file_name: str
    report_file: str
    sheet_name: str
    row_number: int


@dataclass
class _SegmentRef:
    segment_id: str
    source: str
    target: str
    file_path: str


@dataclass
class _MatchOutcome:
    segment: _SegmentRef | None
    reason: str
    matched_by: str


@dataclass
class _FinalIndex:
    by_id: dict[str, list[_SegmentRef]] = field(default_factory=lambda: defaultdict(list))
    by_source_exact: dict[str, list[_SegmentRef]] = field(
        default_factory=lambda: defaultdict(list)
    )
    by_source_norm: dict[str, list[_SegmentRef]] = field(
        default_factory=lambda: defaultdict(list)
    )
    by_source_compact: dict[str, list[_SegmentRef]] = field(
        default_factory=lambda: defaultdict(list)
    )
    by_file_key: dict[str, set[str]] = field(default_factory=lambda: defaultdict(set))
    loaded_files: list[str] = field(default_factory=list)


class QAVerifier:
    _header_aliases = {
        "source_column": [
            "source",
            "source text",
            "src",
            "source sentence",
            "source string",
            "РёСЃС…РѕРґРЅС‹Р№ С‚РµРєСЃС‚",
            "source segment",
            "РёСЃС…РѕРґРЅРёРє",
        ],
        "original_column": [
            "original translation",
            "original",
            "old translation",
            "origin",
            "origin translation",
            "orig translation",
            "current translation",
            "target",
            "translation",
            "initial translation",
            "original target",
            "РёСЃС…РѕРґРЅС‹Р№ РїРµСЂРµРІРѕРґ",
            "РѕСЂРёРіРёРЅР°Р» РїРµСЂРµРІРѕРґР°",
        ],
        "revised_column": [
            "revised translation",
            "corrected translation",
            "fixed translation",
            "suggested translation",
            "new translation",
            "updated translation",
            "qa suggestion",
            "РёСЃРїСЂР°РІР»РµРЅРЅС‹Р№ РїРµСЂРµРІРѕРґ",
            "СЂРµРєРѕРјРµРЅРґРѕРІР°РЅРЅС‹Р№ РїРµСЂРµРІРѕРґ",
        ],
        "qa_mark_column": [
            "qa mark",
            "tp/fp",
            "tp fp",
            "issue type",
            "qa result",
            "qa results",
            "tp/fp result",
            "mark",
            "qa status",
            "qa",
            "С‚РёРї Р·Р°РјРµС‡Р°РЅРёСЏ",
            "РјРµС‚РєР°",
        ],
        "segment_id_column": [
            "segment id",
            "segment",
            "id",
            "tu id",
            "trans-unit id",
            "mid",
            "segment number",
            "РЅРѕРјРµСЂ СЃРµРіРјРµРЅС‚Р°",
        ],
        "filename_column": [
            "file",
            "file name",
            "filename",
            "file_name",
            "document",
            "document name",
            "xliff file",
            "target file",
            "resource",
            "resource name",
        ],
    }

    def __init__(self, on_progress: Callable[[str, float], None] | None = None) -> None:
        self.on_progress = on_progress
        ParserRegistry.discover()

    def scan_reports(self, report_paths: list[str]) -> QAScanResult:
        configs: list[QASheetConfig] = []
        warnings: list[str] = []
        total = len(report_paths) if report_paths else 1

        for index, report_path in enumerate(report_paths, start=1):
            self._progress(
                f"Scanning QA report {index}/{len(report_paths)}...",
                0.05 + (index / total) * 0.45,
            )
            path = Path(report_path)
            if not path.exists():
                warnings.append(f"QA report not found: {report_path}")
                continue
            try:
                workbook = openpyxl.load_workbook(
                    path,
                    read_only=True,
                    data_only=True,
                )
            except Exception as exc:
                warnings.append(f"Failed to read report {report_path}: {exc}")
                continue

            try:
                for sheet_name in workbook.sheetnames:
                    try:
                        worksheet = workbook[sheet_name]
                        config = self._detect_sheet_config(str(path), worksheet)
                        configs.append(config)
                    except Exception as exc:
                        warnings.append(
                            f"Failed to detect mapping for {path.name}::{sheet_name}: {exc}"
                        )
            finally:
                workbook.close()

        return QAScanResult(sheet_configs=configs, warnings=warnings)

    def verify(
        self,
        sheet_configs: list[QASheetConfig],
        final_files: list[str],
    ) -> QAVerificationResult:
        if not sheet_configs:
            return QAVerificationResult(
                rows=[],
                status_counts=self._empty_counts(),
                warnings=["No QA report sheets loaded."],
                total_rows=0,
                timestamp=datetime.now(),
            )

        complete_sheet_configs = [item for item in sheet_configs if item.mapping.is_complete()]
        if not complete_sheet_configs:
            return QAVerificationResult(
                rows=[],
                status_counts=self._empty_counts(),
                warnings=[
                    "No sheets with complete mapping. "
                    "Map Source, Original Translation, and QA mark columns."
                ],
                total_rows=0,
                timestamp=datetime.now(),
            )

        self._progress("Parsing final translation files...", 0.05)
        index, parse_warnings = self._build_final_index(final_files)

        grouped_configs: dict[str, list[QASheetConfig]] = defaultdict(list)
        for config in complete_sheet_configs:
            grouped_configs[config.report_path].append(config)

        all_rows: list[_QARowInput] = []
        warnings: list[str] = []
        skipped_sheets = len(sheet_configs) - len(complete_sheet_configs)
        if skipped_sheets > 0:
            warnings.append(f"Skipped {skipped_sheets} sheet(s) with incomplete mapping.")
        warnings.extend(parse_warnings)

        reports_total = len(grouped_configs) if grouped_configs else 1
        for report_idx, (report_path, configs) in enumerate(grouped_configs.items(), start=1):
            self._progress(
                f"Reading QA rows {report_idx}/{len(grouped_configs)}...",
                0.1 + (report_idx / reports_total) * 0.35,
            )
            try:
                workbook = openpyxl.load_workbook(
                    report_path,
                    read_only=True,
                    data_only=True,
                )
            except Exception as exc:
                warnings.append(f"Failed to read QA report {report_path}: {exc}")
                continue

            try:
                for config in configs:
                    try:
                        worksheet = workbook[config.sheet_name]
                    except Exception:
                        warnings.append(
                            f"Sheet not found: {Path(report_path).name}::{config.sheet_name}"
                        )
                        continue
                    all_rows.extend(self._extract_sheet_rows(worksheet, config))
            finally:
                workbook.close()

        results: list[QAVerificationRow] = []
        total_rows = len(all_rows) if all_rows else 1
        for idx, row in enumerate(all_rows, start=1):
            self._progress(
                f"Verifying QA rows {idx}/{len(all_rows)}...",
                0.5 + (idx / total_rows) * 0.45,
            )
            results.append(self._verify_row(row, index))

        counts = self._empty_counts()
        for item in results:
            counts[item.verification_status] = counts.get(item.verification_status, 0) + 1

        file_summaries = self._build_file_summaries(results, index)
        self._progress("Done", 1.0)
        return QAVerificationResult(
            rows=results,
            status_counts=counts,
            warnings=warnings,
            total_rows=len(results),
            timestamp=datetime.now(),
            file_summaries=file_summaries,
        )

    def export_to_excel(self, result: QAVerificationResult, output_path: str) -> str:
        output_file = Path(output_path)
        if output_file.suffix.lower() != ".xlsx":
            output_file = output_file.with_suffix(".xlsx")
        output_file.parent.mkdir(parents=True, exist_ok=True)

        workbook = xlsxwriter.Workbook(
            str(output_file),
            {
                "strings_to_formulas": False,
                "strings_to_numbers": False,
                "strings_to_urls": False,
            },
        )
        try:
            header_fmt = workbook.add_format({"bold": True, "bg_color": "#f3f4f6"})
            wrap_fmt = workbook.add_format({"text_wrap": True, "valign": "top"})
            diff_formats = {
                "revised_insert": workbook.add_format({"font_color": "#1d4ed8"}),
                "revised_delete": workbook.add_format(
                    {"font_color": "#1d4ed8", "font_strikeout": True}
                ),
                "revised_ws": workbook.add_format(
                    {"font_color": "#1d4ed8", "bg_color": "#dbeafe"}
                ),
                "revised_ws_delete": workbook.add_format(
                    {
                        "font_color": "#1d4ed8",
                        "bg_color": "#dbeafe",
                        "font_strikeout": True,
                    }
                ),
                "final_insert": workbook.add_format({"font_color": "#15803d"}),
                "final_delete": workbook.add_format(
                    {"font_color": "#15803d", "font_strikeout": True}
                ),
                "final_ws": workbook.add_format(
                    {"font_color": "#15803d", "bg_color": "#dcfce7"}
                ),
                "final_ws_delete": workbook.add_format(
                    {
                        "font_color": "#15803d",
                        "bg_color": "#dcfce7",
                        "font_strikeout": True,
                    }
                ),
            }
            status_formats = {
                STATUS_APPLIED: workbook.add_format(
                    {"text_wrap": True, "bg_color": "#ecfdf3", "valign": "top"}
                ),
                STATUS_NOT_APPLIED: workbook.add_format(
                    {"text_wrap": True, "bg_color": "#fef2f2", "valign": "top"}
                ),
                STATUS_CANNOT_VERIFY: workbook.add_format(
                    {"text_wrap": True, "bg_color": "#fffbeb", "valign": "top"}
                ),
                STATUS_NOT_APPLICABLE: workbook.add_format(
                    {"text_wrap": True, "bg_color": "#f8fafc", "valign": "top"}
                ),
            }

            ws = workbook.add_worksheet("Verification")
            headers = [
                "Source",
                "Original Translation",
                "Revised Translation",
                "Final Translation",
                "Expected File",
                "Matched File",
                "QA Mark",
                "Verification Status",
                "Matched Segment ID",
                "Reason / Comment",
                "Report",
                "Sheet",
                "Row",
            ]
            ws.write_row(0, 0, headers, header_fmt)
            ws.freeze_panes(1, 0)
            text_col_width = self._pixels_to_excel_width(250)
            ws.set_column(0, 3, text_col_width)
            ws.set_column(4, 5, 20)
            ws.set_column(6, 6, 10)
            ws.set_column(7, 7, 18)
            ws.set_column(8, 8, 18)
            ws.set_column(9, 9, 44)
            ws.set_column(10, 11, 20)
            ws.set_column(12, 12, 8)

            for row_idx, item in enumerate(result.rows, start=1):
                row_fmt = status_formats.get(item.verification_status, wrap_fmt)
                values = [
                    item.source,
                    item.original_translation,
                    item.revised_translation,
                    item.final_translation,
                    item.expected_file_name,
                    item.matched_file_name,
                    item.qa_mark,
                    item.verification_status,
                    item.matched_segment_id,
                    item.reason,
                    item.report_file,
                    item.sheet_name,
                    item.row_number,
                ]
                for col_idx, value in enumerate(values):
                    if col_idx == 2:
                        self._write_diff_cell(
                            worksheet=ws,
                            row=row_idx,
                            col=col_idx,
                            base_text=item.original_translation,
                            compared_text=item.revised_translation,
                            cell_format=row_fmt,
                            insert_format=diff_formats["revised_insert"],
                            delete_format=diff_formats["revised_delete"],
                            whitespace_insert_format=diff_formats["revised_ws"],
                            whitespace_delete_format=diff_formats["revised_ws_delete"],
                        )
                        continue
                    if col_idx == 3:
                        self._write_diff_cell(
                            worksheet=ws,
                            row=row_idx,
                            col=col_idx,
                            base_text=item.original_translation,
                            compared_text=item.final_translation,
                            cell_format=row_fmt,
                            insert_format=diff_formats["final_insert"],
                            delete_format=diff_formats["final_delete"],
                            whitespace_insert_format=diff_formats["final_ws"],
                            whitespace_delete_format=diff_formats["final_ws_delete"],
                        )
                        continue
                    ws.write_string(row_idx, col_idx, "" if value is None else str(value), row_fmt)
            ws.autofilter(0, 0, max(1, len(result.rows)), len(headers) - 1)

            summary = workbook.add_worksheet("Summary")
            summary.write_row(0, 0, ["Status", "Count"], header_fmt)
            summary_rows = [
                STATUS_APPLIED,
                STATUS_NOT_APPLIED,
                STATUS_CANNOT_VERIFY,
                STATUS_NOT_APPLICABLE,
            ]
            for row_idx, status in enumerate(summary_rows, start=1):
                summary.write_string(row_idx, 0, status)
                summary.write_number(row_idx, 1, result.status_counts.get(status, 0))
            summary.write_string(6, 0, "Total rows")
            summary.write_number(6, 1, result.total_rows)
            summary.write_string(7, 0, "Generated at")
            summary.write_string(7, 1, result.timestamp.strftime("%Y-%m-%d %H:%M:%S"))
            summary.write_row(
                9,
                0,
                [
                    "Translation file",
                    "QA rows",
                    STATUS_APPLIED,
                    STATUS_NOT_APPLIED,
                    STATUS_CANNOT_VERIFY,
                    STATUS_NOT_APPLICABLE,
                    "Info",
                ],
                header_fmt,
            )
            for row_idx, item in enumerate(result.file_summaries, start=10):
                summary.write_string(row_idx, 0, item.file_name)
                summary.write_number(row_idx, 1, item.qa_rows)
                summary.write_number(row_idx, 2, item.applied)
                summary.write_number(row_idx, 3, item.not_applied)
                summary.write_number(row_idx, 4, item.cannot_verify)
                summary.write_number(row_idx, 5, item.not_applicable)
                unresolved = item.not_applied + item.cannot_verify
                if item.qa_rows == 0:
                    info = "No QA issues in report."
                elif unresolved == 0:
                    info = "All QA issues were applied."
                else:
                    info = f"{unresolved} issue(s) require attention."
                summary.write_string(row_idx, 6, info)
            summary.set_column(0, 0, 22)
            summary.set_column(1, 1, 12)
            summary.set_column(2, 5, 14)
            summary.set_column(6, 6, 34)
        finally:
            workbook.close()

        return str(output_file)

    def _detect_sheet_config(self, report_path: str, worksheet) -> QASheetConfig:
        (
            header_row,
            columns,
            mapping,
            required_hits,
            _mapped_hits,
        ) = self._select_best_header_candidate(worksheet)
        notes: list[str] = []
        if header_row > 1:
            notes.append(f"Header row auto-detected at row {header_row}.")
        if not mapping.is_complete():
            missing = [
                self._mapping_field_label(field_name)
                for field_name in REQUIRED_MAPPING_FIELDS
                if not getattr(mapping, field_name)
            ]
            notes.append("Missing required columns: " + ", ".join(missing))
        elif required_hits < len(REQUIRED_MAPPING_FIELDS):
            notes.append("Header detection confidence is low, verify mapping manually.")

        return QASheetConfig(
            report_path=report_path,
            sheet_name=worksheet.title,
            header_row=header_row,
            columns=columns,
            mapping=mapping,
            notes=notes,
        )

    def _select_best_header_candidate(
        self,
        worksheet,
    ) -> tuple[int, list[QAColumnInfo], QAColumnMapping, int, int]:
        max_scan_rows = min(25, max(1, worksheet.max_row or 1))
        best_row = 1
        best_columns = self._extract_columns(worksheet, 1)
        best_mapping = self._detect_mapping(best_columns)
        best_required_hits = sum(
            1 for field_name in REQUIRED_MAPPING_FIELDS if getattr(best_mapping, field_name)
        )
        best_mapped_hits = sum(
            1
            for field_name in self._header_aliases.keys()
            if getattr(best_mapping, field_name)
        )
        best_confidence = self._mapping_confidence(best_columns, best_mapping)
        best_non_empty = self._non_empty_header_count(best_columns)
        best_key = (
            best_required_hits,
            best_mapped_hits,
            best_confidence,
            best_non_empty,
            -best_row,
        )

        for row_index in range(2, max_scan_rows + 1):
            columns = self._extract_columns(worksheet, row_index)
            mapping = self._detect_mapping(columns)
            required_hits = sum(
                1
                for field_name in REQUIRED_MAPPING_FIELDS
                if getattr(mapping, field_name)
            )
            mapped_hits = sum(
                1
                for field_name in self._header_aliases.keys()
                if getattr(mapping, field_name)
            )
            confidence = self._mapping_confidence(columns, mapping)
            non_empty = self._non_empty_header_count(columns)
            candidate_key = (
                required_hits,
                mapped_hits,
                confidence,
                non_empty,
                -row_index,
            )
            if candidate_key > best_key:
                best_row = row_index
                best_columns = columns
                best_mapping = mapping
                best_required_hits = required_hits
                best_mapped_hits = mapped_hits
                best_key = candidate_key

        return (
            best_row,
            best_columns,
            best_mapping,
            best_required_hits,
            best_mapped_hits,
        )

    def _detect_header_row(self, worksheet) -> int:
        max_scan_rows = min(25, max(1, worksheet.max_row or 1))
        max_scan_cols = min(80, max(1, worksheet.max_column or 1))
        best_row = 1
        best_score = -1

        for row_index in range(1, max_scan_rows + 1):
            headers: list[str] = []
            non_empty = 0
            for col_index in range(1, max_scan_cols + 1):
                value = worksheet.cell(row=row_index, column=col_index).value
                text = self._cell_to_text(value)
                if text:
                    non_empty += 1
                headers.append(text)
            if non_empty == 0:
                continue

            field_hits = 0
            total_hit_score = 0
            for aliases in self._header_aliases.values():
                best_hit = 0
                for header in headers:
                    if not header:
                        continue
                    best_hit = max(best_hit, self._header_score(header, aliases))
                if best_hit > 0:
                    field_hits += 1
                    total_hit_score += best_hit

            score = (field_hits * 100) + total_hit_score + non_empty
            if score > best_score:
                best_score = score
                best_row = row_index

        return best_row

    def _extract_columns(self, worksheet, header_row: int) -> list[QAColumnInfo]:
        max_columns = min(150, max(1, worksheet.max_column or 1))
        columns: list[QAColumnInfo] = []
        for col_index in range(1, max_columns + 1):
            letter = get_column_letter(col_index)
            value = worksheet.cell(row=header_row, column=col_index).value
            header = self._cell_to_text(value)
            if not header:
                header = f"Column {letter}"
            columns.append(
                QAColumnInfo(
                    column_index=col_index,
                    column_letter=letter,
                    header=header,
                )
            )
        return columns

    def _detect_mapping(self, columns: list[QAColumnInfo]) -> QAColumnMapping:
        mapping = QAColumnMapping()
        used: set[str] = set()
        ordered_fields = [
            "source_column",
            "original_column",
            "qa_mark_column",
            "revised_column",
            "segment_id_column",
            "filename_column",
        ]
        for field_name in ordered_fields:
            aliases = self._header_aliases[field_name]
            best_column: str | None = None
            best_score = 0
            for item in columns:
                if item.column_letter in used:
                    continue
                score = self._header_score(item.header, aliases)
                if score > best_score:
                    best_score = score
                    best_column = item.column_letter
            if best_column:
                setattr(mapping, field_name, best_column)
                used.add(best_column)
        return mapping

    def _extract_sheet_rows(self, worksheet, config: QASheetConfig) -> list[_QARowInput]:
        if not config.mapping.is_complete():
            return []

        source_idx = self._column_idx(config.mapping.source_column)
        original_idx = self._column_idx(config.mapping.original_column)
        revised_idx = self._column_idx(config.mapping.revised_column)
        qa_idx = self._column_idx(config.mapping.qa_mark_column)
        segment_id_idx = self._column_idx(config.mapping.segment_id_column)
        filename_idx = self._column_idx(config.mapping.filename_column)

        rows: list[_QARowInput] = []
        for row_num, row_values in enumerate(
            worksheet.iter_rows(min_row=config.header_row + 1, values_only=True),
            start=config.header_row + 1,
        ):
            source = self._safe_tuple_value(row_values, source_idx)
            original = self._safe_tuple_value(row_values, original_idx)
            revised = self._safe_tuple_value(row_values, revised_idx)
            qa_mark = self._safe_tuple_value(row_values, qa_idx)
            segment_id = self._safe_tuple_value(row_values, segment_id_idx)
            expected_file_name = self._safe_tuple_value(row_values, filename_idx)

            if not any([source, original, revised, qa_mark, segment_id]):
                continue
            rows.append(
                _QARowInput(
                    source=source,
                    original_translation=original,
                    revised_translation=revised,
                    qa_mark_raw=qa_mark,
                    segment_id=segment_id,
                    expected_file_name=expected_file_name,
                    report_file=Path(config.report_path).name,
                    sheet_name=config.sheet_name,
                    row_number=row_num,
                )
            )
        return rows

    def _build_final_index(self, final_files: list[str]) -> tuple[_FinalIndex, list[str]]:
        index = _FinalIndex()
        warnings: list[str] = []
        total = len(final_files) if final_files else 1

        for file_idx, file_path in enumerate(final_files, start=1):
            self._progress(
                f"Parsing final file {file_idx}/{len(final_files)}...",
                0.05 + (file_idx / total) * 0.25,
            )
            path = Path(file_path)
            ext = path.suffix.lower()
            if ext not in SUPPORTED_FINAL_EXTENSIONS:
                warnings.append(f"Unsupported final file format: {file_path}")
                continue

            try:
                parser = ParserRegistry.get_parser(str(path))
                document = parser.parse(str(path))
            except (UnsupportedFormatError, ParseError) as exc:
                warnings.append(f"Failed to parse final file {file_path}: {exc}")
                continue
            except Exception as exc:
                warnings.append(f"Failed to parse final file {file_path}: {exc}")
                continue

            self._add_document_to_index(index, document)

        return index, warnings

    def _add_document_to_index(self, index: _FinalIndex, document: ParsedDocument) -> None:
        file_path = str(Path(document.file_path))
        if file_path not in index.loaded_files:
            index.loaded_files.append(file_path)
        for key in self._file_lookup_keys(Path(file_path).name):
            index.by_file_key[key].add(file_path)

        for segment in document.segments:
            ref = _SegmentRef(
                segment_id=(segment.id or "").strip(),
                source=segment.source or "",
                target=segment.target or "",
                file_path=file_path,
            )
            if ref.segment_id:
                key_id = self._normalize_segment_id(ref.segment_id)
                index.by_id[key_id].append(ref)

            source = ref.source.strip()
            if not source:
                continue
            exact_key = self._normalize_source_exact(source)
            norm_key = self._normalize_source_norm(source)
            compact_key = self._normalize_source_compact(source)
            if exact_key:
                index.by_source_exact[exact_key].append(ref)
            if norm_key:
                index.by_source_norm[norm_key].append(ref)
            if compact_key:
                index.by_source_compact[compact_key].append(ref)

    def _verify_row(self, row: _QARowInput, final_index: _FinalIndex) -> QAVerificationRow:
        mark = self._normalize_qa_mark(row.qa_mark_raw)
        source = row.source
        original = row.original_translation
        revised = row.revised_translation

        if mark == "FP":
            return QAVerificationRow(
                source=source,
                original_translation=original,
                revised_translation=revised,
                final_translation="",
                expected_file_name=row.expected_file_name,
                matched_file_name="",
                matched_file_path="",
                qa_mark="FP",
                verification_status=STATUS_NOT_APPLICABLE,
                matched_segment_id="",
                reason="FP mark: verification skipped.",
                report_file=row.report_file,
                sheet_name=row.sheet_name,
                row_number=row.row_number,
            )

        if mark != "TP":
            return QAVerificationRow(
                source=source,
                original_translation=original,
                revised_translation=revised,
                final_translation="",
                expected_file_name=row.expected_file_name,
                matched_file_name="",
                matched_file_path="",
                qa_mark=mark or row.qa_mark_raw,
                verification_status=STATUS_CANNOT_VERIFY,
                matched_segment_id="",
                reason="Unsupported QA mark.",
                report_file=row.report_file,
                sheet_name=row.sheet_name,
                row_number=row.row_number,
            )

        if not original:
            return QAVerificationRow(
                source=source,
                original_translation=original,
                revised_translation=revised,
                final_translation="",
                expected_file_name=row.expected_file_name,
                matched_file_name="",
                matched_file_path="",
                qa_mark="TP",
                verification_status=STATUS_CANNOT_VERIFY,
                matched_segment_id="",
                reason="Missing Original Translation.",
                report_file=row.report_file,
                sheet_name=row.sheet_name,
                row_number=row.row_number,
            )

        match = self._match_segment(row, final_index)
        if match.segment is None:
            return QAVerificationRow(
                source=source,
                original_translation=original,
                revised_translation=revised,
                final_translation="",
                expected_file_name=row.expected_file_name,
                matched_file_name="",
                matched_file_path="",
                qa_mark="TP",
                verification_status=STATUS_CANNOT_VERIFY,
                matched_segment_id="",
                reason=match.reason,
                report_file=row.report_file,
                sheet_name=row.sheet_name,
                row_number=row.row_number,
            )

        final_translation = match.segment.target or ""
        original_cmp = self._normalize_for_compare(original)
        final_cmp = self._normalize_for_compare(final_translation)
        if original_cmp == final_cmp:
            status = STATUS_NOT_APPLIED
            reason = f"{match.reason}; Original equals Final."
        else:
            status = STATUS_APPLIED
            reason = f"{match.reason}; Original differs from Final."

        return QAVerificationRow(
            source=source,
            original_translation=original,
            revised_translation=revised,
            final_translation=final_translation,
            expected_file_name=row.expected_file_name,
            matched_file_name=Path(match.segment.file_path).name,
            matched_file_path=match.segment.file_path,
            qa_mark="TP",
            verification_status=status,
            matched_segment_id=match.segment.segment_id,
            reason=reason,
            report_file=row.report_file,
            sheet_name=row.sheet_name,
            row_number=row.row_number,
        )

    def _match_segment(self, row: _QARowInput, final_index: _FinalIndex) -> _MatchOutcome:
        expected_paths: set[str] | None = None
        if row.expected_file_name:
            expected_paths = self._resolve_expected_file_paths(
                row.expected_file_name,
                final_index,
            )
            if not expected_paths:
                return _MatchOutcome(
                    None,
                    f"Referenced file not loaded: {row.expected_file_name}",
                    "file_name",
                )

        if row.segment_id:
            id_key = self._normalize_segment_id(row.segment_id)
            id_candidates = self._filter_candidates_by_files(
                final_index.by_id.get(id_key, []),
                expected_paths,
            )
            selected = self._select_candidate(id_candidates, row.source)
            if selected.segment is not None:
                reason = f"Matched by Segment ID ({Path(selected.segment.file_path).name})"
                return _MatchOutcome(selected.segment, reason, "segment_id")
            if id_candidates:
                return _MatchOutcome(None, "Ambiguous match by Segment ID.", "segment_id")

        if row.source:
            exact_key = self._normalize_source_exact(row.source)
            exact_candidates = self._filter_candidates_by_files(
                final_index.by_source_exact.get(exact_key, []),
                expected_paths,
            )
            selected = self._select_candidate(exact_candidates, row.source)
            if selected.segment is not None:
                reason = f"Matched by exact Source ({Path(selected.segment.file_path).name})"
                return _MatchOutcome(selected.segment, reason, "source_exact")
            if exact_candidates:
                return _MatchOutcome(None, "Ambiguous exact Source match.", "source_exact")

            norm_key = self._normalize_source_norm(row.source)
            norm_candidates = self._filter_candidates_by_files(
                final_index.by_source_norm.get(norm_key, []),
                expected_paths,
            )
            selected = self._select_candidate(norm_candidates, row.source)
            if selected.segment is not None:
                reason = f"Matched by normalized Source ({Path(selected.segment.file_path).name})"
                return _MatchOutcome(selected.segment, reason, "source_normalized")
            if norm_candidates:
                return _MatchOutcome(
                    None,
                    "Ambiguous normalized Source match.",
                    "source_normalized",
                )

            compact_key = self._normalize_source_compact(row.source)
            compact_candidates = self._filter_candidates_by_files(
                final_index.by_source_compact.get(compact_key, []),
                expected_paths,
            )
            selected = self._select_candidate(compact_candidates, row.source)
            if selected.segment is not None:
                reason = f"Matched by compact Source ({Path(selected.segment.file_path).name})"
                return _MatchOutcome(selected.segment, reason, "source_compact")
            if compact_candidates:
                return _MatchOutcome(
                    None,
                    "Ambiguous compact Source match.",
                    "source_compact",
                )

        if expected_paths is not None:
            return _MatchOutcome(None, "Segment not found in referenced file.", "file_name")
        return _MatchOutcome(None, "Segment not found in final files.", "none")

    def _select_candidate(
        self,
        candidates: list[_SegmentRef],
        source_hint: str,
    ) -> _MatchOutcome:
        if not candidates:
            return _MatchOutcome(None, "No candidates.", "none")
        if len(candidates) == 1:
            return _MatchOutcome(candidates[0], "Single candidate.", "unique")

        if source_hint:
            exact_key = self._normalize_source_exact(source_hint)
            exact = [
                item
                for item in candidates
                if self._normalize_source_exact(item.source) == exact_key
            ]
            if len(exact) == 1:
                return _MatchOutcome(exact[0], "Resolved by exact source.", "source_exact")
            if len(exact) > 1:
                candidates = exact
            else:
                norm_key = self._normalize_source_norm(source_hint)
                norm = [
                    item
                    for item in candidates
                    if self._normalize_source_norm(item.source) == norm_key
                ]
                if len(norm) == 1:
                    return _MatchOutcome(norm[0], "Resolved by normalized source.", "source_norm")
                if len(norm) > 1:
                    candidates = norm

        targets = {self._normalize_for_compare(item.target) for item in candidates}
        if len(targets) == 1:
            return _MatchOutcome(candidates[0], "Resolved by identical targets.", "same_target")
        return _MatchOutcome(None, "Multiple conflicting candidate segments.", "ambiguous")

    @staticmethod
    def _filter_candidates_by_files(
        candidates: list[_SegmentRef],
        expected_paths: set[str] | None,
    ) -> list[_SegmentRef]:
        if expected_paths is None:
            return list(candidates)
        return [item for item in candidates if item.file_path in expected_paths]

    def _resolve_expected_file_paths(
        self,
        raw_file_name: str,
        final_index: _FinalIndex,
    ) -> set[str]:
        matches: set[str] = set()
        for key in self._file_lookup_keys(raw_file_name):
            matches.update(final_index.by_file_key.get(key, set()))
        return matches

    @classmethod
    def _file_lookup_keys(cls, raw_file_name: str) -> set[str]:
        normalized = cls._normalize_file_ref(raw_file_name)
        if not normalized:
            return set()
        keys = {normalized}
        stem = Path(normalized).stem.casefold()
        suffix = Path(normalized).suffix.casefold()
        if stem:
            keys.add(stem)
            canonical_stem = cls._strip_copy_suffix(stem)
            keys.add(canonical_stem)
            if suffix:
                keys.add(f"{canonical_stem}{suffix}")
        return keys

    @staticmethod
    def _normalize_file_ref(value: str) -> str:
        text = unicodedata.normalize("NFKC", str(value or ""))
        text = text.strip().strip('"').strip("'")
        if not text:
            return ""
        text = text.replace("\\", "/").split("/")[-1]
        return text.casefold()

    @staticmethod
    def _strip_copy_suffix(stem: str) -> str:
        text = stem.strip()
        if not text:
            return text
        previous = None
        while previous != text:
            previous = text
            text = re.sub(r"\s*\(\d+\)\s*$", "", text).strip()
            text = re.sub(r"\s*\(copy(?:\s*\d+)?\)\s*$", "", text, flags=re.IGNORECASE).strip()
            text = re.sub(r"\s*[-_ ]copy(?:[-_ ]\d+)?\s*$", "", text, flags=re.IGNORECASE).strip()
        return text or stem

    def _build_file_summaries(
        self,
        rows: list[QAVerificationRow],
        final_index: _FinalIndex,
    ) -> list[QAFileSummary]:
        by_path: dict[str, QAFileSummary] = {
            file_path: QAFileSummary(
                file_name=Path(file_path).name,
                file_path=file_path,
            )
            for file_path in final_index.loaded_files
        }

        for row in rows:
            attribution_paths: set[str] = set()
            if row.matched_file_path:
                attribution_paths.add(row.matched_file_path)
            elif row.expected_file_name:
                expected_paths = self._resolve_expected_file_paths(
                    row.expected_file_name,
                    final_index,
                )
                if len(expected_paths) == 1:
                    attribution_paths.update(expected_paths)

            for file_path in attribution_paths:
                summary = by_path.get(file_path)
                if summary is None:
                    summary = QAFileSummary(
                        file_name=Path(file_path).name,
                        file_path=file_path,
                    )
                    by_path[file_path] = summary
                summary.qa_rows += 1
                if row.verification_status == STATUS_APPLIED:
                    summary.applied += 1
                elif row.verification_status == STATUS_NOT_APPLIED:
                    summary.not_applied += 1
                elif row.verification_status == STATUS_CANNOT_VERIFY:
                    summary.cannot_verify += 1
                elif row.verification_status == STATUS_NOT_APPLICABLE:
                    summary.not_applicable += 1

        return sorted(by_path.values(), key=lambda item: item.file_name.casefold())

    def _header_score(self, header: str, aliases: list[str]) -> int:
        normalized = self._normalize_header(header)
        if not normalized:
            return 0
        best = 0
        for alias in aliases:
            alias_norm = self._normalize_header(alias)
            if not alias_norm:
                continue
            if normalized == alias_norm:
                best = max(best, 100)
                continue
            if normalized.startswith(alias_norm) or normalized.endswith(alias_norm):
                best = max(best, 90)
                continue
            if alias_norm in normalized:
                best = max(best, 80)
                continue
            alias_tokens = set(alias_norm.split())
            header_tokens = set(normalized.split())
            common = alias_tokens.intersection(header_tokens)
            if common:
                score = int((len(common) / max(1, len(alias_tokens))) * 60)
                best = max(best, score)
        return best

    def _write_diff_cell(
        self,
        worksheet,
        row: int,
        col: int,
        base_text: str,
        compared_text: str,
        cell_format,
        insert_format,
        delete_format,
        whitespace_insert_format,
        whitespace_delete_format,
    ) -> None:
        base = "" if base_text is None else str(base_text)
        compared = "" if compared_text is None else str(compared_text)
        if not compared:
            worksheet.write_string(row, col, "", cell_format)
            return

        use_symbol_diff = TextDiffer.has_only_non_word_or_case_changes(base, compared)
        chunks = (
            TextDiffer.diff_chars(base, compared)
            if use_symbol_diff
            else TextDiffer.diff_auto(base, compared)
        )
        rich: list[object] = []
        for chunk in chunks:
            if chunk.type == ChunkType.EQUAL:
                self._append_rich_text(rich, chunk.text)
                continue
            if chunk.type == ChunkType.INSERT:
                self._append_rich_text(
                    rich,
                    chunk.text,
                    whitespace_insert_format if chunk.text.strip() == "" else insert_format,
                )
                continue
            if chunk.type == ChunkType.DELETE and use_symbol_diff:
                self._append_rich_text(
                    rich,
                    chunk.text,
                    whitespace_delete_format if chunk.text.strip() == "" else delete_format,
                )

        self._write_rich_or_plain(
            worksheet=worksheet,
            row=row,
            col=col,
            fragments=rich,
            plain_text=compared,
            cell_format=cell_format,
        )

    @staticmethod
    def _append_rich_text(fragments: list[object], text: str, style=None) -> None:
        if not text:
            return
        if style is None:
            if fragments and isinstance(fragments[-1], str):
                fragments[-1] = f"{fragments[-1]}{text}"
            else:
                fragments.append(text)
            return
        fragments.append(style)
        fragments.append(text)

    @staticmethod
    def _write_rich_or_plain(
        worksheet,
        row: int,
        col: int,
        fragments: list[object],
        plain_text: str,
        cell_format,
    ) -> None:
        has_rich_runs = any(not isinstance(item, str) for item in fragments)
        if not has_rich_runs:
            worksheet.write_string(row, col, plain_text, cell_format)
            return
        if fragments and not isinstance(fragments[0], str):
            worksheet.write_string(row, col, plain_text, cell_format)
            return
        if len(fragments) <= 2:
            worksheet.write_string(row, col, plain_text, cell_format)
            return
        try:
            result = worksheet.write_rich_string(row, col, *fragments, cell_format)
            if result != 0:
                worksheet.write_string(row, col, plain_text, cell_format)
        except Exception:
            worksheet.write_string(row, col, plain_text, cell_format)

    @staticmethod
    def _pixels_to_excel_width(px: int) -> float:
        if px <= 12:
            return px / 12.0
        return (px - 5) / 7.0

    def _mapping_confidence(self, columns: list[QAColumnInfo], mapping: QAColumnMapping) -> int:
        by_letter = {item.column_letter: item.header for item in columns}
        score = 0
        for field_name, aliases in self._header_aliases.items():
            selected = getattr(mapping, field_name)
            if not selected:
                continue
            header = by_letter.get(selected, "")
            score += self._header_score(header, aliases)
        return score

    @staticmethod
    def _non_empty_header_count(columns: list[QAColumnInfo]) -> int:
        return sum(
            1
            for item in columns
            if item.header and not item.header.startswith("Column ")
        )

    @staticmethod
    def _normalize_header(value: str) -> str:
        text = unicodedata.normalize("NFKC", str(value or ""))
        text = text.replace("_", " ").replace("-", " ")
        text = re.sub(r"\s+", " ", text).strip().casefold()
        return text

    @staticmethod
    def _mapping_field_label(field_name: str) -> str:
        labels = {
            "source_column": "Source text",
            "original_column": "Original Translation",
            "qa_mark_column": "QA mark (TP/FP)",
        }
        return labels.get(field_name, field_name)

    @staticmethod
    def _cell_to_text(value) -> str:
        if value is None:
            return ""
        return str(value).strip()

    @staticmethod
    def _column_idx(column_letter: str | None) -> int:
        if not column_letter:
            return -1
        return column_index_from_string(column_letter) - 1

    @staticmethod
    def _safe_tuple_value(values, index: int) -> str:
        if index < 0:
            return ""
        if index >= len(values):
            return ""
        value = values[index]
        return "" if value is None else str(value).strip()

    @staticmethod
    def _normalize_qa_mark(value: str) -> str:
        text = (value or "").strip().upper().replace(" ", "")
        if text in {"TP", "TRUEPOSITIVE"}:
            return "TP"
        if text in {"FP", "FALSEPOSITIVE"}:
            return "FP"
        if "TP" in text and "FP" not in text:
            return "TP"
        if "FP" in text and "TP" not in text:
            return "FP"
        return text

    @staticmethod
    def _normalize_segment_id(value: str) -> str:
        return unicodedata.normalize("NFKC", str(value or "")).strip().casefold()

    @classmethod
    def _normalize_for_compare(cls, value: str) -> str:
        text = unicodedata.normalize("NFKC", str(value or ""))
        for char in NBSP_CHARS:
            text = text.replace(char, " ")
        for char in APOSTROPHE_CHARS:
            text = text.replace(char, "'")
        text = text.replace("\r\n", "\n").replace("\r", "\n")
        return text.strip()

    @classmethod
    def _normalize_source_exact(cls, value: str) -> str:
        return cls._normalize_for_compare(value)

    @classmethod
    def _normalize_source_norm(cls, value: str) -> str:
        text = cls._normalize_for_compare(value).casefold()
        text = re.sub(r"\s+", " ", text).strip()
        return text

    @classmethod
    def _normalize_source_compact(cls, value: str) -> str:
        text = cls._normalize_source_norm(value)
        return re.sub(r"[^\w]+", "", text, flags=re.UNICODE)

    @staticmethod
    def _empty_counts() -> dict[str, int]:
        return {
            STATUS_APPLIED: 0,
            STATUS_NOT_APPLIED: 0,
            STATUS_CANNOT_VERIFY: 0,
            STATUS_NOT_APPLICABLE: 0,
        }

    def _progress(self, message: str, value: float) -> None:
        if self.on_progress is not None:
            self.on_progress(message, value)

