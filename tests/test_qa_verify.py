from __future__ import annotations

from pathlib import Path
import zipfile

import openpyxl

from core.qa_verify import (
    QASheetConfig,
    QAVerifier,
    STATUS_APPLIED,
    STATUS_CANNOT_VERIFY,
    STATUS_NOT_APPLICABLE,
    STATUS_NOT_APPLIED,
)


def _write_report(path: Path, rows: list[list[str]]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "QA"
    for row in rows:
        ws.append(row)
    wb.save(path)
    wb.close()


def _write_xliff(path: Path, segments: list[tuple[str, str, str]]) -> None:
    body = []
    for segment_id, source, target in segments:
        body.append(
            "<trans-unit id=\"{id}\"><source>{src}</source>"
            "<target>{tgt}</target></trans-unit>".format(
                id=segment_id,
                src=source.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;"),
                tgt=target.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;"),
            )
        )
    content = (
        "<?xml version=\"1.0\" encoding=\"UTF-8\"?>"
        "<xliff version=\"1.2\"><file original=\"qa\" source-language=\"en\" target-language=\"en\">"
        "<body>{body}</body></file></xliff>"
    ).format(body="".join(body))
    path.write_text(content, encoding="utf-8")


def _first_sheet(scan, report_path: Path) -> QASheetConfig:
    for item in scan.sheet_configs:
        if Path(item.report_path) == report_path:
            return item
    raise AssertionError("sheet config not found")


def test_qa_scan_auto_detects_columns(tmp_path: Path) -> None:
    report = tmp_path / "qa.xlsx"
    _write_report(
        report,
        [
            ["QA Mark", "Original Translation", "Segment ID", "Source Text", "Revised Translation"],
            ["TP", "old", "1", "source", "new"],
        ],
    )
    verifier = QAVerifier()
    scan = verifier.scan_reports([str(report)])
    assert scan.warnings == []
    config = _first_sheet(scan, report)
    assert config.mapping.source_column == "D"
    assert config.mapping.original_column == "B"
    assert config.mapping.qa_mark_column == "A"
    assert config.mapping.segment_id_column == "C"
    assert config.mapping.revised_column == "E"
    assert config.mapping.is_complete()


def test_qa_scan_detects_second_row_headers_and_origin_alias(tmp_path: Path) -> None:
    report = tmp_path / "qa_second_row.xlsx"
    _write_report(
        report,
        [
            ["QA Results", "", "", "", ""],
            ["Source", "Origin", "TP/FP", "Segment ID", "Revised Translation"],
            ["Segment source", "Old target", "TP", "1", "New target"],
        ],
    )
    verifier = QAVerifier()
    scan = verifier.scan_reports([str(report)])
    assert scan.warnings == []
    config = _first_sheet(scan, report)
    assert config.header_row == 2
    assert config.mapping.source_column == "A"
    assert config.mapping.original_column == "B"
    assert config.mapping.qa_mark_column == "C"
    assert config.mapping.segment_id_column == "D"
    assert config.mapping.revised_column == "E"
    assert config.mapping.is_complete()


def test_qa_scan_detects_filename_column_alias(tmp_path: Path) -> None:
    report = tmp_path / "qa_filename.xlsx"
    _write_report(
        report,
        [
            ["FileName", "Source Text", "Original Translation", "QA Mark", "Segment ID"],
            ["f1.xliff", "S", "old", "TP", "1"],
        ],
    )
    verifier = QAVerifier()
    scan = verifier.scan_reports([str(report)])
    config = _first_sheet(scan, report)
    assert config.mapping.filename_column == "A"


def test_qa_verify_uses_original_not_revised_translation(tmp_path: Path) -> None:
    report = tmp_path / "qa.xlsx"
    final_file = tmp_path / "final.xliff"
    _write_report(
        report,
        [
            ["Source Text", "Original Translation", "Revised Translation", "QA Mark", "Segment ID"],
            ["Hello world", "Alpha", "Beta", "TP", "1"],
        ],
    )
    _write_xliff(final_file, [("1", "Hello world", "Alpha")])

    verifier = QAVerifier()
    scan = verifier.scan_reports([str(report)])
    result = verifier.verify(scan.sheet_configs, [str(final_file)])

    assert len(result.rows) == 1
    row = result.rows[0]
    assert row.revised_translation == "Beta"
    assert row.final_translation == "Alpha"
    assert row.verification_status == STATUS_NOT_APPLIED


def test_qa_verify_statuses_tp_fp_and_cannot_verify(tmp_path: Path) -> None:
    report = tmp_path / "qa.xlsx"
    final_file = tmp_path / "final.xliff"
    _write_report(
        report,
        [
            ["Source Text", "Original Translation", "QA Mark", "Segment ID"],
            ["S1", "old", "TP", "1"],
            ["S2", "old2", "FP", "2"],
            ["S3", "old3", "TP", "404"],
        ],
    )
    _write_xliff(final_file, [("1", "S1", "new"), ("2", "S2", "new2")])

    verifier = QAVerifier()
    scan = verifier.scan_reports([str(report)])
    result = verifier.verify(scan.sheet_configs, [str(final_file)])

    assert len(result.rows) == 3
    statuses = [item.verification_status for item in result.rows]
    assert statuses[0] == STATUS_APPLIED
    assert statuses[1] == STATUS_NOT_APPLICABLE
    assert statuses[2] == STATUS_CANNOT_VERIFY
    assert result.status_counts[STATUS_APPLIED] == 1
    assert result.status_counts[STATUS_NOT_APPLICABLE] == 1
    assert result.status_counts[STATUS_CANNOT_VERIFY] == 1


def test_qa_verify_matches_by_normalized_source_and_exports(tmp_path: Path) -> None:
    report = tmp_path / "qa.xlsx"
    final_file = tmp_path / "final.xliff"
    out_file = tmp_path / "verify.xlsx"
    _write_report(
        report,
        [
            ["Source Text", "Original Translation", "QA Mark", "Segment ID"],
            ["Hello, world!", "old", "TP", ""],
        ],
    )
    _write_xliff(final_file, [("11", "hello world", "new")])

    verifier = QAVerifier()
    scan = verifier.scan_reports([str(report)])
    result = verifier.verify(scan.sheet_configs, [str(final_file)])
    assert result.rows[0].verification_status == STATUS_APPLIED
    assert result.rows[0].matched_segment_id == "11"

    exported = verifier.export_to_excel(result, str(out_file))
    assert Path(exported).exists()
    wb = openpyxl.load_workbook(exported, read_only=True)
    assert "Verification" in wb.sheetnames
    assert "Summary" in wb.sheetnames
    wb.close()


def test_qa_export_contains_rich_diff_for_revised_and_final(tmp_path: Path) -> None:
    report = tmp_path / "qa_rich.xlsx"
    final_file = tmp_path / "final.xliff"
    out_file = tmp_path / "verify_rich.xlsx"
    _write_report(
        report,
        [
            [
                "Source Text",
                "Original Translation",
                "Revised Translation",
                "QA Mark",
                "Segment ID",
            ],
            ["S1", "Alpha beta", "Alpha gamma beta", "TP", "1"],
        ],
    )
    _write_xliff(final_file, [("1", "S1", "Alpha delta beta")])

    verifier = QAVerifier()
    scan = verifier.scan_reports([str(report)])
    result = verifier.verify(scan.sheet_configs, [str(final_file)])
    exported = verifier.export_to_excel(result, str(out_file))

    with zipfile.ZipFile(exported) as archive:
        shared_strings = archive.read("xl/sharedStrings.xml").decode("utf-8", errors="ignore")
    assert "FF1D4ED8" in shared_strings.upper()
    assert "FF15803D" in shared_strings.upper()


def test_qa_verify_respects_filename_and_reports_files_without_issues(tmp_path: Path) -> None:
    report = tmp_path / "qa_files.xlsx"
    file_a = tmp_path / "file_a.xliff"
    file_b = tmp_path / "file_b.xliff"
    out_file = tmp_path / "verify_files.xlsx"
    _write_report(
        report,
        [
            ["FileName", "Source Text", "Original Translation", "QA Mark", "Segment ID"],
            ["file_a.xliff", "Source A", "old", "TP", "1"],
        ],
    )
    _write_xliff(file_a, [("1", "Source A", "new")])
    _write_xliff(file_b, [("1", "Source A", "old")])

    verifier = QAVerifier()
    scan = verifier.scan_reports([str(report)])
    result = verifier.verify(scan.sheet_configs, [str(file_a), str(file_b)])

    assert len(result.rows) == 1
    assert result.rows[0].verification_status == STATUS_APPLIED
    assert result.rows[0].matched_file_name == "file_a.xliff"
    by_name = {item.file_name: item for item in result.file_summaries}
    assert by_name["file_a.xliff"].qa_rows == 1
    assert by_name["file_b.xliff"].qa_rows == 0

    exported = verifier.export_to_excel(result, str(out_file))
    wb = openpyxl.load_workbook(exported, read_only=True, data_only=True)
    ws = wb["Summary"]
    values = [cell for row in ws.iter_rows(values_only=True) for cell in row if cell is not None]
    wb.close()
    assert "No QA issues in report." in values


def test_qa_verify_matches_filename_with_copy_suffix(tmp_path: Path) -> None:
    report = tmp_path / "qa_filename_copy.xlsx"
    file_with_copy = (
        tmp_path
        / "seller-portal-dev-content-content-frontend-services-platform-analytics-frontend-feature-atd-7203-uz(ru-uz-Latn) (1).xliff"
    )
    _write_report(
        report,
        [
            ["FileName", "Source Text", "Original Translation", "QA Mark", "Segment ID"],
            [
                "seller-portal-dev-content-content-frontend-services-platform-analytics-frontend-feature-atd-7203-uz(ru-uz-Latn).xliff",
                "Source A",
                "old",
                "TP",
                "1",
            ],
        ],
    )
    _write_xliff(file_with_copy, [("1", "Source A", "new")])

    verifier = QAVerifier()
    scan = verifier.scan_reports([str(report)])
    result = verifier.verify(scan.sheet_configs, [str(file_with_copy)])

    assert len(result.rows) == 1
    row = result.rows[0]
    assert row.verification_status == STATUS_APPLIED
    assert row.matched_file_name == file_with_copy.name
