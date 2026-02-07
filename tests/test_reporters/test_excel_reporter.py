from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path

import openpyxl
import pytest
from openpyxl.utils import column_index_from_string

from core.models import (
    ChangeRecord,
    ChangeStatistics,
    ChangeType,
    ChunkType,
    ComparisonResult,
    DiffChunk,
    ParsedDocument,
    Segment,
    SegmentContext,
)
from reporters.excel_reporter import ExcelReporter


def make_segment(segment_id: str, target: str) -> Segment:
    context = SegmentContext(
        file_path="file.txt",
        location=segment_id,
        position=int(segment_id),
        group=None,
    )
    return Segment(id=segment_id, source=None, target=target, context=context)


def make_doc(name: str, segments: list[Segment]) -> ParsedDocument:
    return ParsedDocument(segments=segments, format_name="TXT", file_path=name)


def test_excel_reporter_generates_file(tmp_path: Path) -> None:
    seg_a = make_segment("1", "Hello world")
    seg_b = make_segment("1", "Hello brave world")
    seg_added = make_segment("2", "New line")
    seg_deleted = make_segment("3", "Old line")

    changes = [
        ChangeRecord(
            type=ChangeType.MODIFIED,
            segment_before=seg_a,
            segment_after=seg_b,
            text_diff=[
                DiffChunk(type=ChunkType.EQUAL, text="Hello "),
                DiffChunk(type=ChunkType.DELETE, text="old "),
                DiffChunk(type=ChunkType.INSERT, text="brave "),
                DiffChunk(type=ChunkType.EQUAL, text="world"),
            ],
            similarity=0.8,
            context=seg_b.context,
        ),
        ChangeRecord(
            type=ChangeType.ADDED,
            segment_before=None,
            segment_after=seg_added,
            text_diff=[],
            similarity=0.0,
            context=seg_added.context,
        ),
        ChangeRecord(
            type=ChangeType.DELETED,
            segment_before=seg_deleted,
            segment_after=None,
            text_diff=[],
            similarity=0.0,
            context=seg_deleted.context,
        ),
    ]

    stats = ChangeStatistics.from_changes(changes)
    result = ComparisonResult(
        file_a=make_doc("a.txt", [seg_a, seg_deleted]),
        file_b=make_doc("b.txt", [seg_b, seg_added]),
        changes=changes,
        statistics=stats,
        timestamp=datetime.now(timezone.utc),
    )

    reporter = ExcelReporter()
    output_file = reporter.generate(result, str(tmp_path / "report.xlsx"))
    output_path = Path(output_file)

    assert output_path.exists()
    assert output_path.stat().st_size > 0

    workbook = openpyxl.load_workbook(output_path, read_only=True)
    assert set(workbook.sheetnames) == {"Report", "Statistics"}

    report_ws = workbook["Report"]
    assert report_ws.max_row == len(changes) + 1
    assert report_ws.max_column == 6

    stats_ws = workbook["Statistics"]
    assert stats_ws.max_row >= 5


def test_excel_reporter_column_widths(tmp_path: Path) -> None:
    result = ComparisonResult(
        file_a=make_doc("a.txt", []),
        file_b=make_doc("b.txt", []),
        changes=[],
        statistics=ChangeStatistics.from_changes([]),
        timestamp=datetime.now(timezone.utc),
    )
    reporter = ExcelReporter()
    output_file = reporter.generate(result, str(tmp_path / "widths.xlsx"))
    workbook = openpyxl.load_workbook(output_file)
    ws = workbook["Report"]

    widths = {
        "A": 6,
        "B": 15,
        "C": 30,
        "D": 45,
        "E": 45,
        "F": 12,
    }

    width_map: dict[int, float] = {}
    for dim in ws.column_dimensions.values():
        if dim.min is None or dim.max is None or dim.width is None:
            continue
        for idx in range(dim.min, dim.max + 1):
            width_map[idx] = dim.width

    for col, expected in widths.items():
        idx = column_index_from_string(col)
        assert idx in width_map
        assert width_map[idx] == pytest.approx(expected, abs=1.0)


def test_excel_reporter_generate_from_json(tmp_path: Path) -> None:
    data = {
        "file_a_name": "a.txt",
        "file_b_name": "b.txt",
        "statistics": {
            "total_segments": 2,
            "added": 0,
            "deleted": 0,
            "modified": 1,
            "unchanged": 1,
            "change_percentage": 0.5,
        },
        "changes": [
            {
                "type": "modified",
                "segment_before": {"id": "1", "source": None, "target": "Hello world"},
                "segment_after": {"id": "1", "source": None, "target": "Hello, world"},
                "text_diff": [
                    {"type": "equal", "text": "Hello"},
                    {"type": "delete", "text": " "},
                    {"type": "insert", "text": ", "},
                    {"type": "equal", "text": "world"},
                ],
            },
            {
                "type": "unchanged",
                "segment_before": {"id": "2", "source": None, "target": "Same"},
                "segment_after": {"id": "2", "source": None, "target": "Same"},
                "text_diff": [],
            },
        ],
    }
    reporter = ExcelReporter()
    output_file = reporter.generate_from_json(data, str(tmp_path / "from_json.xlsx"))
    workbook = openpyxl.load_workbook(output_file, read_only=True)
    report_ws = workbook["Report"]
    assert report_ws.max_row == 3


def test_excel_reporter_old_new_columns_logic(tmp_path: Path) -> None:
    seg_before = make_segment("1", "A old text")
    seg_after = make_segment("1", "A new text")
    seg_added = make_segment("2", "Only new")
    seg_deleted = make_segment("3", "Only old")

    changes = [
        ChangeRecord(
            type=ChangeType.MODIFIED,
            segment_before=seg_before,
            segment_after=seg_after,
            text_diff=[
                DiffChunk(type=ChunkType.EQUAL, text="A "),
                DiffChunk(type=ChunkType.DELETE, text="old "),
                DiffChunk(type=ChunkType.INSERT, text="new "),
                DiffChunk(type=ChunkType.EQUAL, text="text"),
            ],
            similarity=0.9,
            context=seg_after.context,
        ),
        ChangeRecord(
            type=ChangeType.ADDED,
            segment_before=None,
            segment_after=seg_added,
            text_diff=[],
            similarity=0.0,
            context=seg_added.context,
        ),
        ChangeRecord(
            type=ChangeType.DELETED,
            segment_before=seg_deleted,
            segment_after=None,
            text_diff=[],
            similarity=0.0,
            context=seg_deleted.context,
        ),
    ]
    result = ComparisonResult(
        file_a=make_doc("a.txt", [seg_before, seg_deleted]),
        file_b=make_doc("b.txt", [seg_after, seg_added]),
        changes=changes,
        statistics=ChangeStatistics.from_changes(changes),
        timestamp=datetime.now(timezone.utc),
    )

    reporter = ExcelReporter()
    output_file = reporter.generate(result, str(tmp_path / "logic.xlsx"))
    workbook = openpyxl.load_workbook(output_file)
    ws = workbook["Report"]

    # Row 2: MODIFIED
    assert ws["D2"].value == "A old text"
    assert ws["E2"].value == "A old new text"
    # Row 3: ADDED
    assert ws["D3"].value in ("", None)
    assert ws["E3"].value == "Only new"
    # Row 4: DELETED
    assert ws["D4"].value == "Only old"
    assert ws["E4"].value in ("", None)


def test_excel_reporter_hides_unchanged_by_default(tmp_path: Path) -> None:
    seg = make_segment("1", "Same text")
    changes = [
        ChangeRecord(
            type=ChangeType.UNCHANGED,
            segment_before=seg,
            segment_after=seg,
            text_diff=[],
            similarity=1.0,
            context=seg.context,
        )
    ]
    result = ComparisonResult(
        file_a=make_doc("a.txt", [seg]),
        file_b=make_doc("b.txt", [seg]),
        changes=changes,
        statistics=ChangeStatistics.from_changes(changes),
        timestamp=datetime.now(timezone.utc),
    )

    reporter = ExcelReporter()
    output_file = reporter.generate(result, str(tmp_path / "hidden.xlsx"))
    workbook = openpyxl.load_workbook(output_file)
    ws = workbook["Report"]
    assert ws.auto_filter.ref == "A1:F2"
    assert ws.row_dimensions[2].hidden is True
