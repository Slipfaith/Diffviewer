from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path

import openpyxl

from core.models import (
    BatchFileResult,
    BatchResult,
    ChangeRecord,
    ChangeStatistics,
    ChangeType,
    ComparisonResult,
    MultiVersionResult,
    ParsedDocument,
    Segment,
    SegmentContext,
)
from reporters.summary_reporter import SummaryReporter


def _segment(segment_id: str, source: str, target: str, position: int) -> Segment:
    return Segment(
        id=segment_id,
        source=source,
        target=target,
        context=SegmentContext(
            file_path="test.xliff",
            location=segment_id,
            position=position,
            group=None,
        ),
    )


def _doc(path: str, segments: list[Segment]) -> ParsedDocument:
    return ParsedDocument(segments=segments, format_name="XLIFF", file_path=path)


def test_summary_reporter_marks_non_text_changes_symbol_level(tmp_path: Path) -> None:
    doc1 = _doc("v1.xliff", [_segment("1", "Greeting", "Hello, world", 1)])
    doc2 = _doc("v2.xliff", [_segment("1", "Greeting", "Hello world", 1)])
    doc3 = _doc("v3.xliff", [_segment("1", "Greeting", "Hello  world", 1)])

    result = MultiVersionResult(
        file_paths=["v1.xliff", "v2.xliff", "v3.xliff"],
        comparisons=[],
        documents=[doc1, doc2, doc3],
    )
    output = SummaryReporter().generate_versions(result, str(tmp_path / "versions_summary.html"))
    text = Path(output).read_text(encoding="utf-8")

    assert "version-del-1" in text
    assert "symbol-del" in text
    assert "ws-change" in text
    assert "data-filter=\"changed\"" in text


def test_summary_reporter_fills_targets_by_source_when_ids_change(tmp_path: Path) -> None:
    doc1 = _doc(
        "v1.sdlxliff",
        [
            _segment("1", "Source one.", "Target one v1", 1),
            _segment("2", "Source two?", "Target two v1", 2),
        ],
    )
    doc2 = _doc(
        "v2.sdlxliff",
        [
            _segment("A-101", "Source one.", "Target one v2", 1),
            _segment("A-202", "Source two?", "Target two v2", 2),
        ],
    )
    doc3 = _doc(
        "v3.sdlxliff",
        [
            _segment("B-501", "Source one", "Target one v3", 1),
            _segment("B-502", "Source two", "Target two v3", 2),
        ],
    )

    result = MultiVersionResult(
        file_paths=["v1.sdlxliff", "v2.sdlxliff", "v3.sdlxliff"],
        comparisons=[],
        documents=[doc1, doc2, doc3],
    )
    output = SummaryReporter().generate_versions(result, str(tmp_path / "versions_summary.html"))
    text = Path(output).read_text(encoding="utf-8")
    plain = re.sub(r"<[^>]+>", "", text)

    assert "Target one v2" in plain
    assert "Target one v3" in plain
    assert "Target two v2" in plain
    assert "Target two v3" in plain
    assert "<td class=\"state-missing\"></td>" not in text


def test_summary_reporter_batch_excel_all_changes_has_no_type_column(tmp_path: Path) -> None:
    seg_before = _segment("1", "Shared source", "Old target", 1)
    seg_after = _segment("2", "Shared source", "New target", 1)
    comparison = ComparisonResult(
        file_a=_doc("a.xliff", [seg_before]),
        file_b=_doc("b.xliff", [seg_after]),
        changes=[
            ChangeRecord(
                type=ChangeType.MODIFIED,
                segment_before=seg_before,
                segment_after=seg_after,
                text_diff=[],
                similarity=0.0,
                context=seg_after.context,
            )
        ],
        statistics=ChangeStatistics(
            total_segments=1,
            added=0,
            deleted=0,
            modified=1,
            moved=0,
            unchanged=0,
            change_percentage=1.0,
        ),
        timestamp=datetime.now(),
    )
    batch = BatchResult(
        folder_a="a",
        folder_b="b",
        files=[
            BatchFileResult(
                filename="file.xliff",
                status="compared",
                report_paths=[],
                statistics=comparison.statistics,
                comparison=comparison,
            )
        ],
    )

    output = SummaryReporter().generate_batch_excel(batch, str(tmp_path / "batch.xlsx"))
    workbook = openpyxl.load_workbook(output)
    ws = workbook["All Changes"]

    headers = [cell.value for cell in ws[1]]
    assert headers == ["File", "Segment ID", "Source", "Old Target", "New Target"]
    assert ws.max_row == 2
    assert ws["C2"].value == "Shared source"
    assert ws["D2"].value == "Old target"
    assert ws["E2"].value == "New target"
