from __future__ import annotations

from pathlib import Path

import openpyxl

from core.models import ParsedDocument, ParseError, Segment, SegmentContext
from parsers.base import BaseParser


def _parse_column_reference(column: str | int | None) -> int | None:
    if column is None:
        return None
    if isinstance(column, int):
        if column < 1:
            raise ValueError("Excel source column must be >= 1.")
        return column

    raw = str(column).strip()
    if not raw:
        return None
    if raw.isdigit():
        value = int(raw)
        if value < 1:
            raise ValueError("Excel source column must be >= 1.")
        return value
    if raw.isalpha():
        value = 0
        for char in raw.upper():
            value = value * 26 + (ord(char) - 64)
        return value
    raise ValueError(f"Invalid Excel source column: {column}")


class XlsxParser(BaseParser):
    name = "XLSX Parser"
    supported_extensions = [".xlsx"]
    format_description = "Excel Workbook"

    def __init__(self) -> None:
        self._source_column_index: int | None = None

    def can_handle(self, filepath: str) -> bool:
        return Path(filepath).suffix.lower() in self.supported_extensions

    def set_source_column(self, column: str | int | None) -> None:
        self._source_column_index = _parse_column_reference(column)

    def parse(self, filepath: str) -> ParsedDocument:
        try:
            workbook = openpyxl.load_workbook(filepath, data_only=True)
        except Exception as exc:
            raise ParseError(filepath, str(exc)) from exc

        segments: list[Segment] = []
        source_column_index = self._source_column_index
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                source_value = self._read_row_source_value(row, source_column_index)
                for cell in row:
                    value = cell.value
                    if value is None or value == "":
                        continue
                    if source_column_index is not None and cell.column == source_column_index:
                        continue
                    cell_id = f"{sheet.title}!{cell.coordinate}"
                    context = SegmentContext(
                        file_path=filepath,
                        location=cell_id,
                        position=len(segments) + 1,
                        group=sheet.title,
                    )
                    metadata = {
                        "row": cell.row,
                        "column": cell.column,
                        "sheet_name": sheet.title,
                    }
                    if source_column_index is not None:
                        metadata["source_column"] = source_column_index
                    segments.append(
                        Segment(
                            id=cell_id,
                            source=source_value,
                            target=str(value),
                            context=context,
                            metadata=metadata,
                        )
                    )

        return ParsedDocument(
            segments=segments,
            format_name=self.format_description,
            file_path=filepath,
            metadata={},
            encoding=None,
        )

    def validate(self, filepath: str) -> list[str]:
        errors: list[str] = []
        try:
            _ = openpyxl.load_workbook(filepath, data_only=True)
        except Exception as exc:
            errors.append(str(exc))
        return errors

    @staticmethod
    def _read_row_source_value(row, source_column_index: int | None) -> str | None:
        if source_column_index is None:
            return None
        source_cell_position = source_column_index - 1
        if source_cell_position < 0 or source_cell_position >= len(row):
            return None
        value = row[source_cell_position].value
        if value is None:
            return None
        text = str(value)
        return text if text != "" else None
