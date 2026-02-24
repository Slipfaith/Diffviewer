from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.utils import column_index_from_string
import xlsxwriter

from core.diff_engine import TextDiffer
from core.models import ChunkType, DiffChunk


def _normalize(value: object) -> str | None:
    """Return str or None (treat empty string same as None)."""
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


def _ensure_spaces_between_colored(chunks: list[DiffChunk]) -> list[DiffChunk]:
    """
    Insert a normal-space chunk between adjacent DELETE/INSERT chunks that have
    no whitespace at their boundary.  TextDiffer is designed for two-column
    display; without this pass, consecutive coloured spans in a single cell
    can appear merged (e.g. '~~старое~~новое' instead of '~~старое~~ новое').
    """
    result: list[DiffChunk] = []
    for i, chunk in enumerate(chunks):
        result.append(chunk)
        if i < len(chunks) - 1:
            nxt = chunks[i + 1]
            if (
                chunk.type != ChunkType.EQUAL
                and nxt.type != ChunkType.EQUAL
                and not chunk.text.endswith(" ")
                and not nxt.text.startswith(" ")
            ):
                result.append(DiffChunk(type=ChunkType.EQUAL, text=" "))
    return result


def _build_rich_fragments(chunks: list[DiffChunk], fmt_del, fmt_ins) -> list:
    """
    Convert DiffChunk list into a flat list suitable for xlsxwriter
    write_rich_string().  Pattern:
      EQUAL  → accumulated into a plain string (no format prefix)
      DELETE → fmt_del, text
      INSERT → fmt_ins, text
    """
    chunks = _ensure_spaces_between_colored(chunks)
    fragments: list = []
    text_buf: list[str] = []

    for chunk in chunks:
        if not chunk.text:
            continue
        if chunk.type == ChunkType.EQUAL:
            text_buf.append(chunk.text)
        else:
            if text_buf:
                fragments.append("".join(text_buf))
                text_buf.clear()
            fmt = fmt_del if chunk.type == ChunkType.DELETE else fmt_ins
            fragments.append(fmt)
            fragments.append(chunk.text)

    if text_buf:
        fragments.append("".join(text_buf))

    return fragments


class XlsxColumnReporter:
    """
    Generates a column-by-column comparison Excel report using xlsxwriter.

    Reads both source workbooks with openpyxl (data_only), diffs each cell
    pair with TextDiffer (the same engine used by the HTML / standard Excel
    reporters), then writes the result with xlsxwriter's write_rich_string:

      red strikethrough : text in file_a removed in file_b
      blue              : text added in file_b
      black (normal)    : unchanged text
    """

    def generate(self, file_a: str, file_b: str, output_path: str) -> str:
        wb_a = openpyxl.load_workbook(file_a, data_only=True)
        wb_b = openpyxl.load_workbook(file_b, data_only=True)

        Path(output_path).parent.mkdir(parents=True, exist_ok=True)

        workbook = xlsxwriter.Workbook(
            output_path,
            {"strings_to_formulas": False, "strings_to_urls": False},
        )

        fmt_cell = workbook.add_format({"text_wrap": True})
        fmt_del = workbook.add_format({"font_color": "#FF0000", "font_strikeout": True})
        fmt_ins = workbook.add_format({"font_color": "#0070C0"})

        for ws_a, ws_b in zip(wb_a.worksheets, wb_b.worksheets):
            ws_out = workbook.add_worksheet(ws_b.title)
            self._fill_sheet(ws_a, ws_b, ws_out, fmt_cell, fmt_del, fmt_ins)
            self._copy_dimensions(ws_b, ws_out)

        workbook.close()
        return output_path

    # ------------------------------------------------------------------
    # Private helpers
    # ------------------------------------------------------------------

    def _fill_sheet(self, ws_a, ws_b, ws_out, fmt_cell, fmt_del, fmt_ins) -> None:
        max_row = max(ws_a.max_row or 1, ws_b.max_row or 1)
        max_col = max(ws_a.max_column or 1, ws_b.max_column or 1)

        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                val_a = _normalize(ws_a.cell(row=row, column=col).value)
                val_b = _normalize(ws_b.cell(row=row, column=col).value)

                r, c = row - 1, col - 1  # xlsxwriter uses 0-based indices

                if val_a is None and val_b is None:
                    pass  # leave blank
                elif val_a is None:
                    # Cell only in file_b — show as blue
                    ws_out.write_rich_string(r, c, fmt_ins, val_b, fmt_cell)
                elif val_b is None:
                    # Cell only in file_a — show as red strikethrough
                    ws_out.write_rich_string(r, c, fmt_del, val_a, fmt_cell)
                elif val_a == val_b:
                    ws_out.write(r, c, val_b, fmt_cell)
                else:
                    self._write_diff(ws_out, r, c, val_a, val_b, fmt_cell, fmt_del, fmt_ins)

    def _write_diff(
        self, ws_out, row: int, col: int,
        text_a: str, text_b: str,
        fmt_cell, fmt_del, fmt_ins,
    ) -> None:
        chunks = TextDiffer.diff_auto(text_a, text_b)
        has_changes = any(c.type != ChunkType.EQUAL for c in chunks)
        if not has_changes:
            ws_out.write(row, col, text_b, fmt_cell)
            return

        fragments = _build_rich_fragments(chunks, fmt_del, fmt_ins)
        has_rich = any(not isinstance(f, str) for f in fragments)

        # write_rich_string needs at least one format+text pair
        if not has_rich or len(fragments) <= 2:
            ws_out.write(row, col, text_b, fmt_cell)
            return

        try:
            ret = ws_out.write_rich_string(row, col, *fragments, fmt_cell)
            if ret != 0:
                ws_out.write(row, col, text_b, fmt_cell)
        except Exception:
            ws_out.write(row, col, text_b, fmt_cell)

    def _copy_dimensions(self, ws_src, ws_out) -> None:
        for col_letter, col_dim in ws_src.column_dimensions.items():
            if col_dim.width:
                col_idx = column_index_from_string(col_letter) - 1
                ws_out.set_column(col_idx, col_idx, col_dim.width)
        for row_num, row_dim in ws_src.row_dimensions.items():
            if row_dim.height:
                ws_out.set_row(row_num - 1, row_dim.height)
